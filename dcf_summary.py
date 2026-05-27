"""
Wingstop DCF Model — generate summary charts from Excel workbook.
Outputs: dcf_fcf_projections.png, dcf_valuation_bridge.png, dcf_sensitivity.png
"""

import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np

wb = openpyxl.load_workbook("Wingstop DCF Model.xlsx", data_only=True)

# ── Shared style ──────────────────────────────────────────────────────────────
NAVY   = "#0F3460"
RED    = "#E94560"
GRAY   = "#6B7280"
LIGHT  = "#F3F4F6"
YEARS  = ["FY2025E", "FY2026E", "FY2027E", "FY2028E", "FY2029E"]

plt.rcParams.update({
    "font.family": "DejaVu Sans",
    "axes.spines.top": False,
    "axes.spines.right": False,
    "axes.grid": True,
    "grid.color": "#E5E7EB",
    "grid.linewidth": 0.5,
    "figure.facecolor": "white",
    "axes.facecolor": "white",
})

# ── 1. FCF Projections bar chart ──────────────────────────────────────────────
ws_fcf = wb["FCF Build"]
rows = list(ws_fcf.iter_rows(values_only=True))

revenue_row = next(r for r in wb["Income Statement"].iter_rows(values_only=True)
                   if r[1] == "Total Revenue")
revenue     = [revenue_row[i] for i in range(6, 11)]
ufcf        = [rows[8][i] for i in range(2, 7)]   # UFCF row
nopat       = [rows[4][i] for i in range(2, 7)]

fig, axes = plt.subplots(1, 2, figsize=(11, 4.5))
fig.suptitle("Wingstop (WING) — Projected Financials", fontsize=13,
             fontweight="bold", color=NAVY, y=1.01)

# Revenue
axes[0].bar(YEARS, revenue, color=NAVY, alpha=0.85, width=0.55)
for i, v in enumerate(revenue):
    axes[0].text(i, v + 4, f"${v:.0f}mm", ha="center", va="bottom",
                 fontsize=9, color=NAVY, fontweight="bold")
axes[0].set_title("Total Revenue  ($mm)", fontsize=11, color=NAVY, pad=8)
axes[0].set_ylim(0, max(revenue) * 1.18)
axes[0].tick_params(axis="x", labelsize=8.5)
axes[0].yaxis.set_visible(False)
axes[0].spines["left"].set_visible(False)

# UFCF
axes[1].bar(YEARS, ufcf, color=RED, alpha=0.85, width=0.55)
for i, v in enumerate(ufcf):
    axes[1].text(i, v + 1.5, f"${v:.1f}mm", ha="center", va="bottom",
                 fontsize=9, color=RED, fontweight="bold")
axes[1].set_title("Unlevered Free Cash Flow  ($mm)", fontsize=11, color=NAVY, pad=8)
axes[1].set_ylim(0, max(ufcf) * 1.2)
axes[1].tick_params(axis="x", labelsize=8.5)
axes[1].yaxis.set_visible(False)
axes[1].spines["left"].set_visible(False)

fig.tight_layout()
fig.savefig("dcf_fcf_projections.png", dpi=150, bbox_inches="tight")
plt.close()
print("Saved dcf_fcf_projections.png")

# ── 2. Valuation bridge ───────────────────────────────────────────────────────
ws_dcf = wb["DCF Valuation"]
dcf_rows = list(ws_dcf.iter_rows(values_only=True))

pv_fcf      = dcf_rows[9][2]
pv_tv       = dcf_rows[16][2]
ev          = dcf_rows[22][2]
net_dbt     = dcf_rows[23][2]
eq_val      = dcf_rows[24][2]
shares      = dcf_rows[25][2]
implied_raw = dcf_rows[26][2]
implied     = implied_raw / 1000   # correct units: $mm equity / mm shares = $/share

labels = ["PV of FCFs", "PV of\nTerminal Value", "Enterprise\nValue",
          "( – ) Net Debt", "Equity\nValue"]
values = [pv_fcf, pv_tv, ev, -net_dbt, eq_val]
colors = [NAVY, NAVY, "#1D4ED8", RED, "#059669"]

fig, ax = plt.subplots(figsize=(10, 5))
fig.suptitle("Wingstop DCF — Valuation Bridge  ($mm)", fontsize=13,
             fontweight="bold", color=NAVY)

bar_starts = [0, pv_fcf, 0, ev, ev - net_dbt]
bar_heights = [pv_fcf, pv_tv, ev, -net_dbt, eq_val]

for i, (start, height, label, color) in enumerate(
        zip(bar_starts, bar_heights, labels, colors)):
    if i in (2, 4):   # total bars start from 0
        ax.bar(i, abs(height), bottom=0, color=color, alpha=0.9, width=0.55,
               zorder=3)
    else:
        ax.bar(i, abs(height), bottom=min(start, start + height),
               color=color, alpha=0.85, width=0.55, zorder=3)
    top = (0 if i in (2, 4) else min(start, start + height)) + abs(height)
    ax.text(i, top + 30, f"${abs(height):.0f}mm", ha="center", va="bottom",
            fontsize=9, fontweight="bold",
            color=color if i != 3 else RED)

ax.set_xticks(range(5))
ax.set_xticklabels(labels, fontsize=9.5)
ax.yaxis.set_visible(False)
ax.spines["left"].set_visible(False)
ax.set_ylim(0, max(ev, eq_val) * 1.18)

# Implied price callout
ax.annotate(
    f"Implied Share Price\n${implied:.2f}  (corrected from model)",
    xy=(4, eq_val), xytext=(3.3, eq_val * 0.55),
    arrowprops=dict(arrowstyle="->", color=GRAY, lw=1.2),
    fontsize=9.5, color="#059669", fontweight="bold",
    bbox=dict(boxstyle="round,pad=0.3", facecolor="white",
              edgecolor="#059669", linewidth=1)
)

fig.tight_layout()
fig.savefig("dcf_valuation_bridge.png", dpi=150, bbox_inches="tight")
plt.close()
print("Saved dcf_valuation_bridge.png")

# ── 3. Sensitivity heatmap ─────────────────────────────────────────────────────
ws_sens = wb["Sensitivity"]
sens_rows = list(ws_sens.iter_rows(values_only=True))

# Table 1 — WACC × TGR (rows 3-9 in sheet, index 2-8)
tgr_labels  = [f"{v*100:.1f}%" for v in [0.02, 0.025, 0.03, 0.035, 0.04]]
data_rows   = [r for r in sens_rows[4:11] if isinstance(r[1], float)]
wacc_labels = [f"{r[1]*100:.1f}%" for r in data_rows]
table1      = np.array([[r[i] / 1000 for i in range(2, 7)] for r in data_rows])

fig, ax = plt.subplots(figsize=(8, 4.5))
fig.suptitle("Sensitivity: Implied Share Price  ($)  —  WACC × Terminal Growth Rate",
             fontsize=11, fontweight="bold", color=NAVY)

im = ax.imshow(table1, cmap="RdYlGn", aspect="auto")

ax.set_xticks(range(5))
ax.set_xticklabels(tgr_labels, fontsize=9)
ax.set_yticks(range(len(wacc_labels)))
ax.set_yticklabels(wacc_labels, fontsize=9)
ax.set_xlabel("Terminal Growth Rate", fontsize=10, color=GRAY)
ax.set_ylabel("WACC", fontsize=10, color=GRAY)

# Base case highlight
base_row = wacc_labels.index("7.8%") if "7.8%" in wacc_labels else 1
base_col = tgr_labels.index("3.0%") if "3.0%" in tgr_labels else 2
ax.add_patch(mpatches.Rectangle(
    (base_col - 0.5, base_row - 0.5), 1, 1,
    fill=False, edgecolor=NAVY, linewidth=2.5, zorder=5
))

for i in range(table1.shape[0]):
    for j in range(table1.shape[1]):
        ax.text(j, i, f"${table1[i, j]:.0f}",
                ha="center", va="center", fontsize=8.5,
                color="black", fontweight="bold" if (i == base_row and j == base_col) else "normal")

fig.colorbar(im, ax=ax, label="Implied Share Price ($)", shrink=0.85)
fig.tight_layout()
fig.savefig("dcf_sensitivity.png", dpi=150, bbox_inches="tight")
plt.close()
print("Saved dcf_sensitivity.png")
print("\nAll charts generated successfully.")
